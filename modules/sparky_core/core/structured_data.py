from __future__ import annotations

import csv
import io
import json
import zipfile
from typing import Any, Dict, Iterable, List, Tuple
from xml.etree import ElementTree
from xml.sax.saxutils import escape

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - optional dependency
    Workbook = None
    load_workbook = None


class _DefaultDialect(csv.Dialect):
    delimiter = ","
    quotechar = '"'
    escapechar = None
    doublequote = True
    skipinitialspace = False
    lineterminator = "\n"
    quoting = csv.QUOTE_MINIMAL


def _sniff_dialect(raw_text: str) -> csv.Dialect:
    sample = raw_text[:2048]
    try:
        return csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t", "|"])
    except csv.Error:
        return _DefaultDialect()


def _has_header(raw_text: str, dialect: csv.Dialect) -> bool:
    sample = raw_text[:2048]
    try:
        return csv.Sniffer().has_header(sample)
    except csv.Error:
        return False


def _clean_header(header: List[str]) -> List[str]:
    cleaned: List[str] = []
    seen = set()
    for idx, value in enumerate(header, start=1):
        name = value.strip() or f"col_{idx}"
        if name in seen:
            suffix = 2
            while f"{name}_{suffix}" in seen:
                suffix += 1
            name = f"{name}_{suffix}"
        seen.add(name)
        cleaned.append(name)
    return cleaned


def _collect_columns(rows: Iterable[Dict[str, Any]]) -> List[str]:
    columns: List[str] = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                columns.append(key)
    return columns


def _looks_like_xlsx(filename: str | None, content_type: str | None) -> bool:
    if filename:
        name = filename.lower()
        if name.endswith(".xlsx") or name.endswith(".xls"):
            return True
    if content_type:
        lowered = content_type.lower()
        if "spreadsheetml" in lowered or "ms-excel" in lowered:
            return True
    return False


def _column_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha())
    index = 0
    for ch in letters:
        index = index * 26 + (ord(ch.upper()) - 64)
    return index


def _column_letter(index: int) -> str:
    letters = ""
    while index > 0:
        index, rem = divmod(index - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _read_shared_strings(archive: zipfile.ZipFile) -> List[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    xml_data = archive.read("xl/sharedStrings.xml")
    root = ElementTree.fromstring(xml_data)
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    strings: List[str] = []
    for item in root.findall(f"{ns}si"):
        parts = []
        for node in item.findall(f".//{ns}t"):
            if node.text:
                parts.append(node.text)
        strings.append("".join(parts))
    return strings


def _select_sheet_path(archive: zipfile.ZipFile) -> str | None:
    if "xl/worksheets/sheet1.xml" in archive.namelist():
        return "xl/worksheets/sheet1.xml"
    candidates = [
        name
        for name in archive.namelist()
        if name.startswith("xl/worksheets/") and name.endswith(".xml")
    ]
    return sorted(candidates)[0] if candidates else None


def _looks_like_header_row(values: List[str]) -> bool:
    cleaned = [value.strip() for value in values if value is not None and str(value).strip()]
    if not cleaned:
        return False
    has_alpha = any(any(ch.isalpha() for ch in value) for value in cleaned)
    unique = len({value.lower() for value in cleaned}) == len(cleaned)
    return has_alpha and unique


def _parse_xlsx_openpyxl(
    raw_bytes: bytes,
) -> Tuple[List[Dict[str, Any]], List[str], Dict[str, Any], str | None]:
    if load_workbook is None:
        return [], [], {}, "openpyxl not available."
    try:
        workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception:
        return [], [], {}, "Invalid XLSX file."

    try:
        if not workbook.worksheets:
            return [], [], {}, "XLSX does not contain any worksheets."
        sheet = workbook.worksheets[0]
        max_cols = sheet.max_column or 0
        if max_cols == 0:
            return [], [], {}, "XLSX is empty."

        rows_list: List[List[Any]] = []
        for row in sheet.iter_rows(max_col=max_cols, values_only=True):
            row_values = ["" if cell is None else str(cell) for cell in row]
            if any(str(value).strip() for value in row_values):
                rows_list.append(row_values)

        if not rows_list:
            return [], [], {}, "XLSX is empty."

        header_values = [str(value) if value is not None else "" for value in rows_list[0]]
        header_detected = _looks_like_header_row(header_values)
        if header_detected:
            header = _clean_header(header_values)
            data_rows = rows_list[1:]
        else:
            header = [f"col_{idx}" for idx in range(1, max_cols + 1)]
            data_rows = rows_list

        rows: List[Dict[str, Any]] = []
        for row in data_rows:
            row_dict = {header[idx]: row[idx] if idx < len(row) else "" for idx in range(len(header))}
            rows.append(row_dict)

        report = {
            "format": "xlsx",
            "row_count": len(rows),
            "column_count": len(header),
            "header_detected": header_detected,
            "sheet": sheet.title,
        }
        return rows, header, report, None
    finally:
        workbook.close()


def parse_xlsx_bytes(
    raw_bytes: bytes,
) -> Tuple[List[Dict[str, Any]], List[str], Dict[str, Any], str | None]:
    if load_workbook is not None:
        rows, header, report, error = _parse_xlsx_openpyxl(raw_bytes)
        if error is None:
            return rows, header, report, None
    try:
        with zipfile.ZipFile(io.BytesIO(raw_bytes)) as archive:
            sheet_path = _select_sheet_path(archive)
            if not sheet_path:
                return [], [], {}, "XLSX does not contain any worksheets."

            shared_strings = _read_shared_strings(archive)
            xml_data = archive.read(sheet_path)
    except zipfile.BadZipFile:
        return [], [], {}, "Invalid XLSX file."

    root = ElementTree.fromstring(xml_data)
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    sheet_data = root.find(f"{ns}sheetData")
    if sheet_data is None:
        return [], [], {}, "XLSX worksheet is empty."

    row_dicts: List[Dict[int, Any]] = []
    max_cols = 0
    for row in sheet_data.findall(f"{ns}row"):
        cell_map: Dict[int, Any] = {}
        for idx, cell in enumerate(row.findall(f"{ns}c"), start=1):
            cell_ref = cell.get("r")
            col_index = _column_index(cell_ref) if cell_ref else idx
            cell_type = cell.get("t")
            value = ""
            if cell_type == "s":
                raw_value = cell.findtext(f"{ns}v")
                if raw_value is not None:
                    try:
                        shared_index = int(raw_value)
                        value = shared_strings[shared_index] if shared_index < len(shared_strings) else ""
                    except ValueError:
                        value = ""
            elif cell_type == "inlineStr":
                parts = []
                for node in cell.findall(f".//{ns}t"):
                    if node.text:
                        parts.append(node.text)
                value = "".join(parts)
            elif cell_type == "b":
                value = "TRUE" if cell.findtext(f"{ns}v") == "1" else "FALSE"
            else:
                raw_value = cell.findtext(f"{ns}v")
                value = raw_value if raw_value is not None else ""
            if value is None:
                value = ""
            cell_map[col_index] = value
            if col_index > max_cols:
                max_cols = col_index
        if cell_map:
            row_dicts.append(cell_map)

    if not row_dicts:
        return [], [], {}, "XLSX is empty."

    rows_list: List[List[Any]] = []
    for row_map in row_dicts:
        row_values = [row_map.get(idx, "") for idx in range(1, max_cols + 1)]
        if any(str(value).strip() for value in row_values):
            rows_list.append(row_values)

    if not rows_list:
        return [], [], {}, "XLSX is empty."

    header_values = [str(value) if value is not None else "" for value in rows_list[0]]
    header_detected = _looks_like_header_row(header_values)
    if header_detected:
        header = _clean_header(header_values)
        data_rows = rows_list[1:]
    else:
        header = [f"col_{idx}" for idx in range(1, max_cols + 1)]
        data_rows = rows_list

    rows: List[Dict[str, Any]] = []
    for row in data_rows:
        row_dict = {header[idx]: row[idx] if idx < len(row) else "" for idx in range(len(header))}
        rows.append(row_dict)

    report = {
        "format": "xlsx",
        "row_count": len(rows),
        "column_count": len(header),
        "header_detected": header_detected,
        "sheet": sheet_path.split("/")[-1].replace(".xml", ""),
    }
    return rows, header, report, None


def _normalize_rows(data: Any) -> Tuple[List[Dict[str, Any]], List[str], str | None]:
    if isinstance(data, dict) and "data" in data and isinstance(data["data"], list):
        data = data["data"]

    if isinstance(data, dict):
        columns = list(data.keys())
        return [data], columns, None

    if isinstance(data, list):
        if not data:
            return [], [], "No rows found."
        if all(isinstance(item, dict) for item in data):
            columns = _collect_columns(data)
            return list(data), columns, None
        if all(isinstance(item, list) for item in data):
            max_len = max(len(item) for item in data)
            columns = [f"col_{idx}" for idx in range(1, max_len + 1)]
            rows: List[Dict[str, Any]] = []
            for item in data:
                row = {columns[idx]: item[idx] if idx < len(item) else "" for idx in range(max_len)}
                rows.append(row)
            return rows, columns, None

    return [], [], "Unsupported JSON shape."


def parse_json_text(raw_text: str) -> Tuple[List[Dict[str, Any]], List[str], Dict[str, Any], str | None]:
    try:
        data = json.loads(raw_text)
    except json.JSONDecodeError as exc:
        return [], [], {}, f"Invalid JSON: {exc.msg}."
    rows, columns, error = _normalize_rows(data)
    report = {"format": "json", "row_count": len(rows), "column_count": len(columns)}
    return rows, columns, report, error


def parse_csv_text(raw_text: str) -> Tuple[List[Dict[str, Any]], List[str], Dict[str, Any], str | None]:
    if not raw_text.strip():
        return [], [], {}, "CSV is empty."
    dialect = _sniff_dialect(raw_text)
    header_detected = _has_header(raw_text, dialect)
    reader = csv.reader(io.StringIO(raw_text), dialect=dialect)

    rows: List[Dict[str, Any]] = []
    header: List[str] | None = None
    for row in reader:
        if not row or not any(cell.strip() for cell in row):
            continue
        if header is None:
            if header_detected:
                header = _clean_header(row)
                continue
            header = [f"col_{idx}" for idx in range(1, len(row) + 1)]
        row_dict = {header[idx]: row[idx] if idx < len(row) else "" for idx in range(len(header))}
        rows.append(row_dict)

    if header is None:
        return [], [], {}, "CSV is empty."

    report = {
        "format": "csv",
        "row_count": len(rows),
        "column_count": len(header),
        "delimiter": getattr(dialect, "delimiter", ","),
        "header_detected": header_detected,
    }
    return rows, header, report, None


def detect_format(
    raw_text: str,
    *,
    filename: str | None = None,
    content_type: str | None = None,
) -> str:
    if content_type and "json" in content_type.lower():
        return "json"
    if content_type and "csv" in content_type.lower():
        return "csv"
    if content_type and ("spreadsheetml" in content_type.lower() or "ms-excel" in content_type.lower()):
        return "xlsx"
    if filename:
        name = filename.lower()
        if name.endswith(".json"):
            return "json"
        if name.endswith(".csv"):
            return "csv"
        if name.endswith(".tsv"):
            return "csv"
        if name.endswith(".xlsx") or name.endswith(".xls"):
            return "xlsx"
    stripped = raw_text.lstrip()
    if stripped.startswith("{") or stripped.startswith("["):
        return "json"
    return "csv"


def parse_structured_text(
    raw_text: str | bytes,
    *,
    filename: str | None = None,
    content_type: str | None = None,
) -> Tuple[List[Dict[str, Any]], List[str], Dict[str, Any], str, str | None]:
    if isinstance(raw_text, bytes) and _looks_like_xlsx(filename, content_type):
        rows, columns, report, error = parse_xlsx_bytes(raw_text)
        return rows, columns, report, "xlsx", error

    text = raw_text.decode("utf-8", errors="replace") if isinstance(raw_text, bytes) else raw_text
    fmt = detect_format(text, filename=filename, content_type=content_type)
    if fmt == "json":
        rows, columns, report, error = parse_json_text(text)
        return rows, columns, report, fmt, error
    if fmt == "csv":
        rows, columns, report, error = parse_csv_text(text)
        return rows, columns, report, fmt, error
    if fmt == "xlsx":
        if isinstance(raw_text, bytes):
            rows, columns, report, error = parse_xlsx_bytes(raw_text)
            return rows, columns, report, fmt, error
        return [], [], {"format": "xlsx"}, fmt, "XLSX upload required."
    return [], [], {"format": fmt}, fmt, "Unsupported format."


def rows_to_csv(rows: List[Dict[str, Any]], columns: List[str]) -> str:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=columns, lineterminator="\n")
    writer.writeheader()
    for row in rows:
        writer.writerow({col: row.get(col, "") for col in columns})
    return output.getvalue()


def rows_to_json(rows: List[Dict[str, Any]]) -> str:
    return json.dumps(rows, ensure_ascii=True, indent=2)


def rows_to_xlsx_bytes(rows: List[Dict[str, Any]], columns: List[str]) -> bytes:
    if Workbook is not None:
        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "data"
        sheet.append(columns)
        for row in rows:
            sheet.append([row.get(column, "") for column in columns])
        workbook.save(output)
        return output.getvalue()

    sheet_rows = [columns]
    for row in rows:
        sheet_rows.append([row.get(column, "") for column in columns])

    row_xml_chunks = []
    for row_idx, values in enumerate(sheet_rows, start=1):
        cell_chunks = []
        for col_idx, value in enumerate(values, start=1):
            text = str(value) if value is not None else ""
            if not text.strip():
                continue
            cell_ref = f"{_column_letter(col_idx)}{row_idx}"
            safe = escape(text)
            cell_chunks.append(
                f'<c r="{cell_ref}" t="inlineStr"><is><t>{safe}</t></is></c>'
            )
        row_xml = f'<row r="{row_idx}">{"".join(cell_chunks)}</row>'
        row_xml_chunks.append(row_xml)

    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        "<sheetData>"
        f"{''.join(row_xml_chunks)}"
        "</sheetData>"
        "</worksheet>"
    )

    content_types = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        "</Types>"
    )

    rels = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="xl/workbook.xml"/>'
        "</Relationships>"
    )

    workbook = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        "<sheets>"
        '<sheet name="data" sheetId="1" r:id="rId1"/>'
        "</sheets>"
        "</workbook>"
    )

    workbook_rels = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
        'Target="worksheets/sheet1.xml"/>'
        "</Relationships>"
    )

    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types)
        archive.writestr("_rels/.rels", rels)
        archive.writestr("xl/workbook.xml", workbook)
        archive.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
        archive.writestr("xl/worksheets/sheet1.xml", sheet_xml)
    return output.getvalue()


def profile_rows(rows: List[Dict[str, Any]], columns: List[str]) -> Dict[str, Any]:
    row_count = len(rows)
    col_count = len(columns)
    empty_rows = 0
    missing_cells = 0
    column_stats: Dict[str, Dict[str, Any]] = {}
    for column in columns:
        column_stats[column] = {"missing": 0, "filled": 0, "samples": []}

    for row in rows:
        row_empty = True
        for column in columns:
            value = row.get(column)
            if value is None or str(value).strip() == "":
                missing_cells += 1
                column_stats[column]["missing"] += 1
            else:
                row_empty = False
                column_stats[column]["filled"] += 1
                samples = column_stats[column]["samples"]
                if len(samples) < 3:
                    samples.append(str(value))
        if row_empty:
            empty_rows += 1

    for column in columns:
        filled = column_stats[column]["filled"]
        column_stats[column]["fill_rate"] = filled / row_count if row_count else 0

    return {
        "row_count": row_count,
        "column_count": col_count,
        "empty_rows": empty_rows,
        "missing_cells": missing_cells,
        "columns": column_stats,
    }


def row_signatures(rows: List[Dict[str, Any]], columns: List[str]) -> List[Tuple[str, ...]]:
    signatures = []
    for row in rows:
        signature = tuple(str(row.get(col, "")).strip() for col in columns)
        signatures.append(signature)
    return signatures
